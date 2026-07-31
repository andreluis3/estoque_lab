import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ==============================
# Caminhos
# ==============================

SERVIDOR = r"I:\LGE\operacao\Areas\CEM\software\EstoqueLab"

LOCAL = r"C:\Users\Public\Documents\EstoqueLab"


PASTA_EXPORT_SERVIDOR = os.path.join(
    SERVIDOR,
    "planilha_equipamentos"
)


PASTA_EXPORT_LOCAL = os.path.join(
    LOCAL,
    "planilha_equipamentos"
)


# ==============================
# Serviço de Exportação
# ==============================

class ExportService:


    def __init__(
        self,
        item_repository,
        historico_repository
    ):

        self.item_repository = item_repository
        self.historico_repository = historico_repository


    # ==============================
    # Verifica caminho disponível
    # ==============================
    def _obter_pasta_exportacao(self):

            """
            Prioridade:
            1 - Servidor
            2 - Máquina local
            """

            if os.path.exists(SERVIDOR):

                os.makedirs(
                    PASTA_EXPORT_SERVIDOR,
                    exist_ok=True
                )

                return PASTA_EXPORT_SERVIDOR


            os.makedirs(
                PASTA_EXPORT_LOCAL,
                exist_ok=True
            )

            return PASTA_EXPORT_LOCAL



    # ==============================
    # Exportar Estoque
    # ==============================

    def exportar_estoque_excel(self):


        itens = self.item_repository.listar()


        pasta = self._obter_pasta_exportacao()


        caminho = os.path.join(
            pasta,
            "estoque_atual.xlsx"
        )


        workbook = Workbook()


        sheet = workbook.active

        sheet.title = "Estoque"



        cabecalho = [
            "ID",
            "Nome",
            "Tipo",
            "Modelo",
            "Quantidade",
            "Caixa",
            "Localização",
            "Slot"
        ]


        sheet.append(cabecalho)



        for coluna in sheet[1]:

            coluna.font = Font(
                bold=True
            )

            coluna.alignment = Alignment(
                horizontal="center"
            )



        for item in itens:


            sheet.append(
                [
                    item["id"],
                    item["nome"],
                    item["tipo"],
                    item["modelo"],
                    item["quantidade"],
                    item["caixa"],
                    item["localizacao"],
                    item["slot"]
                ]
            )



        workbook.save(caminho)


        return caminho





    def exportar_historico_excel(self):
        historico = self.historico_repository.listar()
        pasta = self._obter_pasta_exportacao()
        caminho = os.path.join(
            pasta,
            "historico.xlsx"
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Histórico"



        cabecalho = [
            "ID",
            "Item",
            "Operação",
            "Quantidade Anterior",
            "Quantidade Nova",
            "Usuário",
            "Data"
        ]

        sheet.append(cabecalho)

        for coluna in sheet[1]:

            coluna.font = Font(
                bold=True
            )



        for registro in historico:

            sheet.append(
                [
                    registro["id"],
                    registro["item_nome"],
                    registro["acao"],
                    registro["quantidade_anterior"],
                    registro["quantidade_nova"],
                    registro["usuario"],
                    registro["data"]
                ]
            )



        workbook.save(caminho)


        return caminho
