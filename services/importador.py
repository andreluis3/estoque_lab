"""
importador.py — Importação limpa da planilha Excel para o banco SQLite.

ESTRATÉGIA DE SYNC:
  - Usa UPSERT (INSERT OR REPLACE) baseado no índice único (nome, modelo)
  - Quantidade vem DIRETAMENTE da planilha (fonte da verdade)
  - NÃO soma, NÃO acumula — substitui o valor correto
  - Registra log de cada item inserido/atualizado
"""

import pandas as pd
from database.db import conectar_db
from openpyxl import load_workbook
from controllers.crud import Crud
import os
from typing import Optional
COLUNAS_PLANILHA = {
    "Nome Item": "nome",
    "Tipo": "tipo",
    "Modelo": "modelo",
    "Quantidade": "quantidade",
    "Caixa": "caixa",
    "Localização": "localizacao",
    "Slot": "slot"
}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CAMINHO_PADRAO = os.path.join(BASE_DIR, "..", "planilhas", "estoque_lab_formatada.xlsx")

def importar_excel(caminho: str) -> list[dict]:
    df = pd.read_excel(caminho, header=1)
    
    
    if "Nome Item" not in df.columns:
        df = pd.read_excel(caminho, header=1)

    print("📊 COLUNAS:", df.columns.tolist())
    print("📊 LINHAS BRUTAS:", len(df))

    df.columns = df.columns.str.strip()
    df = df.rename(columns=COLUNAS_PLANILHA)

    for col in COLUNAS_PLANILHA.values():
        if col not in df.columns:
            df[col] = None

    df["nome"] = df["nome"].fillna("").astype(str).str.strip()
    df["tipo"] = df["tipo"].fillna("Outros").astype(str).str.strip()
    df["modelo"] = df["modelo"].fillna("").astype(str).str.strip()
    df["quantidade"] = pd.to_numeric(df["quantidade"], errors="coerce").fillna(0).astype(int)
    df["caixa"] = df["caixa"].fillna("").astype(str).str.strip()
    df["localizacao"] = df["localizacao"].fillna("Não informado").astype(str).str.strip()
    df["slot"] = df["slot"].fillna("Não informado").astype(str).str.strip()

    df = df[df["nome"] != ""]

    print("📊 LINHAS VÁLIDAS:", len(df))

    df = df.drop_duplicates(subset=["nome", "modelo"], keep="last")

    return df.to_dict(orient="records")


def importar_para_banco(caminho: Optional[str] = None, usuario: str = "importador", resetar: bool = False) -> dict:
    if not caminho:
        caminho = CAMINHO_PADRAO

    print(f"📥 IMPORTANDO PLANILHA: {caminho}")
    print(f"📂 Caminho absoluto: {os.path.abspath(caminho)}")
    print(f"📄 Arquivo existe? {os.path.exists(caminho)}")
    print(f"IMPORTANDO PLANILHA: {caminho}")
    """
    Importa a planilha para o banco de dados.

    Args:
        caminho: Caminho do arquivo .xlsx
        usuario: Quem disparou a importação (para o log)
        resetar: Se True, APAGA o banco antes de importar (use com cuidado!)
                 Se False (padrão), faz UPSERT seguro — preserva histórico.

    Returns:
        dict com estatísticas da importação
    """
    conn = conectar_db()
    cursor = conn.cursor()

    dados = importar_excel(caminho)

    if resetar:
        cursor.execute("DELETE FROM movimentacoes")
        cursor.execute("DELETE FROM historico_alteracoes")
        cursor.execute("DELETE FROM itens")
        cursor.execute("DELETE FROM sqlite_sequence WHERE name='itens'")
        conn.commit()
        _registrar_auditoria_global(cursor, usuario, "reset_banco",
                                    f"Banco resetado antes de importar {len(dados)} itens")

    inseridos = 0
    atualizados = 0
    erros = []

    for item in dados:
        try:
            # Verifica se já existe no banco
            existente = cursor.execute("""
                SELECT id, quantidade, tipo, caixa, localizacao, slot
                FROM itens WHERE nome = ? AND modelo = ?
            """, (item["nome"], item["modelo"])).fetchone()

            if existente:
                item_id = existente[0]
                qtd_antiga = existente[1]

                # Atualiza com os valores corretos da planilha (não soma!)
                cursor.execute("""
                    UPDATE itens
                    SET tipo=?, quantidade=?, caixa=?, localizacao=?, slot=?,
                        atualizado_em=CURRENT_TIMESTAMP
                    WHERE id=?
                """, (
                    item["tipo"], item["quantidade"], item["caixa"],
                    item["localizacao"], item["slot"], item_id
                ))

                # Loga cada campo que mudou
                campos_para_verificar = {
                    "quantidade": (str(qtd_antiga), str(item["quantidade"])),
                    "tipo":       (existente[2], item["tipo"]),
                    "caixa":      (existente[3], item["caixa"]),
                    "localizacao":(existente[4], item["localizacao"]),
                    "slot":       (existente[5], item["slot"]),
                }
                for campo, (antes, depois) in campos_para_verificar.items():
                    if antes != depois:
                        _registrar_historico(cursor, item_id, campo, antes, depois,
                                             usuario, "atualizado_importacao")

                # Movimentação de quantidade se mudou
                diff = item["quantidade"] - qtd_antiga
                if diff != 0:
                    tipo_mov = "entrada" if diff > 0 else "saida"
                    cursor.execute("""
                        INSERT INTO movimentacoes (item_id, tipo, quantidade, usuario)
                        VALUES (?, ?, ?, ?)
                    """, (item_id, tipo_mov, abs(diff), usuario))

                atualizados += 1

            else:
                cursor.execute("""
                    INSERT INTO itens (nome, tipo, modelo, quantidade, caixa, localizacao, slot)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    item["nome"], item["tipo"], item["modelo"], item["quantidade"],
                    item["caixa"], item["localizacao"], item["slot"]
                ))
                item_id = cursor.lastrowid

                cursor.execute("""
                    INSERT INTO movimentacoes (item_id, tipo, quantidade, usuario)
                    VALUES (?, ?, ?, ?)
                """, (item_id, "entrada", item["quantidade"], usuario))

                _registrar_historico(cursor, item_id, "*", None,
                                     f"{item['nome']} | {item['modelo']}",
                                     usuario, "inserido_importacao")
                inseridos += 1

        except Exception as e:
            erros.append({"item": item.get("nome"), "erro": str(e)})

    conn.commit()
    conn.close()

    resultado = {
        "status": "ok",
        "total_planilha": len(dados),
        "inseridos": inseridos,
        "atualizados": atualizados,
        "erros": erros
    }

    print(f"✅ Importação concluída: {inseridos} inseridos, {atualizados} atualizados, {len(erros)} erros")
    if erros:
        for e in erros:
            print(f"  ⚠️  {e['item']}: {e['erro']}")

    return resultado


# ─── helpers internos ────────────────────────────────────────────────────────

def _registrar_historico(cursor, item_id, campo, valor_anterior, valor_novo, usuario, acao):
    cursor.execute("""
        INSERT INTO historico_alteracoes (item_id, campo, valor_anterior, valor_novo, usuario, acao)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (item_id, campo, valor_anterior, valor_novo, usuario, acao))


def _registrar_auditoria_global(cursor, usuario, acao, descricao):
    cursor.execute("""
        INSERT INTO historico_alteracoes (item_id, campo, valor_anterior, valor_novo, usuario, acao)
        VALUES (NULL, 'sistema', NULL, ?, ?, ?)
    """, (descricao, usuario, acao))


def salvar_com_template(dados: list[dict], caminho_template: str, caminho_saida: str) -> None:
    """
    Salva os dados numa cópia do template Excel, preservando a formatação original.

    Args:
        dados:            lista de dicts com os itens do banco
        caminho_template: caminho do .xlsx original (ex: "planilhas/estoque_lab_formatada.xlsx")
        caminho_saida:    onde salvar o arquivo gerado  (ex: "planilhas/exportado.xlsx")

    Quem chama esta função passa os caminhos — ela não os adivinha.
    Exemplo de uso no janela_principal.py:
        from services.importador import salvar_com_template
        itens = self.service.listar_itens()
        salvar_com_template(itens, "planilhas/estoque_lab_completa.xlsx", "planilhas/exportado.xlsx")
    """
    wb = load_workbook(caminho_template)

    # wb.active pode ser None se o workbook não tiver nenhuma aba ativa
    ws = wb.active
    if ws is None:
        raise ValueError(f"O arquivo '{caminho_template}' não possui nenhuma aba ativa.")

    # Apaga linhas de dados (mantém o cabeçalho na linha 1)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    for i, item in enumerate(dados, start=2):
        ws[f"A{i}"] = item["nome"]
        ws[f"B{i}"] = item["tipo"]
        ws[f"C{i}"] = item["modelo"]
        ws[f"D{i}"] = item["quantidade"]
        ws[f"E{i}"] = item["caixa"]
        ws[f"F{i}"] = item["localizacao"]
        ws[f"G{i}"] = item["slot"]

    wb.save(caminho_saida)
    
def gerar_relatorio_inconsistencias(caminho: str) -> list[dict]:
        """
        Lê a planilha e retorna itens com problemas:
        nome vazio, quantidade negativa, duplicatas, etc.
        """
        dados = importar_excel(caminho)
        problemas = []

        vistos = {}
        for item in dados:
            chave = (item["nome"], item["modelo"])

            if not item["nome"]:
                problemas.append({"item": item, "problema": "Nome vazio"})

            if item["quantidade"] < 0:
                problemas.append({"item": item, "problema": "Quantidade negativa"})

            if chave in vistos:
                problemas.append({"item": item, "problema": f"Duplicata de '{item['nome']} | {item['modelo']}'"})
            else:
                vistos[chave] = True

        return problemas
    
def validar_diferencas(caminho_excel):
    """
    Compara quantidades da planilha Excel com o banco de dados.
    Retorna lista de itens com discrepâncias.
    """
    crud = Crud()
    df_excel = pd.read_excel(caminho_excel)
    df_excel.columns = df_excel.columns.str.strip()
    df_excel = df_excel.rename(columns={
        "Nome Item": "nome",
        "Tipo": "tipo",
        "Modelo": "modelo",
        "Quantidade": "quantidade",
        "Caixa": "caixa",
        "Localização": "localizacao",
        "Slot": "slot"
    })

    banco = crud.listar_itens()
    mapa_banco = {
        (item["nome"].strip().lower(), item["modelo"].strip().lower()): item
        for item in banco
    }

    erros = []

    for _, row in df_excel.iterrows():
        chave = (
            str(row.get("nome", "")).strip().lower(),
            str(row.get("modelo", "")).strip().lower()
        )
        qtd_excel = int(row.get("quantidade", 0))

        if chave in mapa_banco:
            qtd_banco = mapa_banco[chave]["quantidade"]
            if qtd_excel != qtd_banco:
                erros.append({
                    "nome": chave[0],
                    "modelo": chave[1],
                    "excel": qtd_excel,
                    "banco": qtd_banco
                })
        else:
            erros.append({
                "nome": chave[0],
                "modelo": chave[1],
                "excel": qtd_excel,
                "banco": "NÃO EXISTE"
            })

    return erros
    
    
